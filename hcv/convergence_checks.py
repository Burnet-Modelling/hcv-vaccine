"""
Convergence check — HCV prevalence only.

Figures:
  1. 3x2 time series grid (n=100,250,500,750,1000 + overlay)
     Each panel: median + shaded 95% CI across sampled runs
  2. Point estimate figure — median at 2026 with 95% CI error bars,
     one point per sample size

Excel:
  - Time-series table (year x sample-size: median, lower, upper)
  - 2026 summary table (sample size x scenario: median, lower, upper, CI width)
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import sciris as sc
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sample_sizes = [100, 250, 500, 750, 1000]
year_start   = 2000   # show full time series for context
year_end     = 2050
reference_year   = 2026
outcome_key  = "HCV prevalence"

colors_list = {
    100:  "#E63946",
    250:  "#F4A261",
    500:  "#2A9D8F",
    750:  "#457B9D",
    1000: "#1D3557",
}


# ---------------------------------------------------------------------------
# Core computation
# ---------------------------------------------------------------------------

def get_run_cols(df):
    return [c for c in df.columns if c.startswith("run_")]


def sample_stats(df: pd.DataFrame, n: int) -> pd.DataFrame:
    """
    For a given outcome DataFrame, sample n run columns and compute
    per-year median + 95% CI across those runs.
    Returns DataFrame with columns: year, median, lower, upper, ci_width
    """
    run_cols = get_run_cols(df)
    sampled = run_cols[:n]

    mask = (df["year"] >= year_start - 0.5) & (df["year"] <= year_end + 0.5)
    df_f = df.loc[mask, ["year"] + sampled].copy()

    result = pd.DataFrame()
    result["year"]     = df_f["year"].values.astype(float)
    result["median"]   = df_f[sampled].median(axis=1).values.astype(float)
    result["lower"]    = df_f[sampled].quantile(0.025, axis=1).values.astype(float)
    result["upper"]    = df_f[sampled].quantile(0.975, axis=1).values.astype(float)
    result["ci_width"] = result["upper"] - result["lower"]
    return result


def focal_year_stats(stats_df: pd.DataFrame) -> dict:
    """Extract stats at the row closest to reference_year."""
    idx = (stats_df["year"] - reference_year).abs().idxmin()
    row = stats_df.loc[idx]
    return {
        "median":   float(row["median"]),
        "lower":    float(row["lower"]),
        "upper":    float(row["upper"]),
        "ci_width": float(row["ci_width"]),
    }


# ---------------------------------------------------------------------------
# Figure 1 — 3x2 time series grid
# ---------------------------------------------------------------------------

def plot_timeseries_grid(all_stats: dict, scenario_key: str, country: str) -> plt.Figure:
    """
    all_stats: {n: stats_df}
    3x2 grid: 5 individual panels + 1 overlay panel
    """
    fig = plt.figure(figsize=(15, 10))
    fig.suptitle(
        f"{country}  |  {outcome_key}  |  Scenario: {scenario_key}\n"
        f"Median + 95% CI across sampled runs",
        fontsize=13, fontweight="bold", y=0.98
    )
    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.42, wspace=0.32)

    # shared y-limits across individual panels
    y_min = min(s["lower"].min() for s in all_stats.values()) * 0.95
    y_max = max(s["upper"].max() for s in all_stats.values()) * 1.05

    positions = [(0,0),(0,1),(0,2),(1,0),(1,1)]

    for pos, n in zip(positions, sample_sizes):
        ax = fig.add_subplot(gs[pos])
        st = all_stats[n]
        col = colors_list[n]

        ax.fill_between(st["year"], st["lower"], st["upper"],
                        color=col, alpha=0.25, label="95% CI")
        ax.plot(st["year"], st["median"],
                color=col, linewidth=1.8, label="Median")
        ax.axvline(reference_year, color="black", linestyle="--",
                   linewidth=0.9, alpha=0.6)

        ax.set_title(f"n = {n:,} runs", fontsize=11, fontweight="bold", color=col)
        ax.set_xlabel("Year", fontsize=9)
        ax.set_ylabel("Prevalence", fontsize=9)
        ax.set_ylim(y_min, y_max)
        ax.set_xlim(year_start, year_end)
        ax.tick_params(labelsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
        ax.grid(True, alpha=0.3, linewidth=0.5)
        ax.spines[["top","right"]].set_visible(False)

    # Overlay panel (bottom-right)
    ax_ov = fig.add_subplot(gs[1, 2])
    for n in sample_sizes:
        st = all_stats[n]
        col = colors_list[n]
        ax_ov.fill_between(st["year"], st["lower"], st["upper"],
                           color=col, alpha=0.12)
        ax_ov.plot(st["year"], st["median"],
                   color=col, linewidth=1.4, label=f"n={n:,}")
    ax_ov.axvline(reference_year, color="black", linestyle="--",
                  linewidth=0.9, alpha=0.6)
    ax_ov.set_title("All sample sizes", fontsize=11, fontweight="bold", color="#333333")
    ax_ov.set_xlabel("Year", fontsize=9)
    ax_ov.set_ylabel("Prevalence", fontsize=9)
    ax_ov.set_xlim(year_start, year_end)
    ax_ov.tick_params(labelsize=8)
    ax_ov.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    ax_ov.grid(True, alpha=0.3, linewidth=0.5)
    ax_ov.spines[["top","right"]].set_visible(False)
    ax_ov.legend(fontsize=8, framealpha=0.7)

    return fig


# ---------------------------------------------------------------------------
# Figure 2 — Median + CI at focal year across sample sizes
# ---------------------------------------------------------------------------

def plot_focal_year(all_focal: dict, scenario_key: str, country: str) -> plt.Figure:
    """
    all_focal: {n: {"median":..., "lower":..., "upper":..., "ci_width":...}}
    Two-panel figure:
      Left  — median with error bars (absolute values)
      Right — CI width vs sample size
    """
    ns      = list(all_focal.keys())
    medians = [all_focal[n]["median"]   for n in ns]
    lowers  = [all_focal[n]["lower"]    for n in ns]
    uppers  = [all_focal[n]["upper"]    for n in ns]
    widths  = [all_focal[n]["ci_width"] for n in ns]
    colors  = [colors_list[n] for n in ns]

    yerr_lower = [m - l for m, l in zip(medians, lowers)]
    yerr_upper = [u - m for m, u in zip(medians, uppers)]

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    fig.suptitle(
        f"{country}  |  {outcome_key} at {reference_year}  |  Scenario: {scenario_key}\n"
        f"Convergence of median and 95% CI across sample sizes",
        fontsize=12, fontweight="bold", y=1.01
    )

    # --- Left: median + CI error bars ---
    ax = axes[0]
    for i, n in enumerate(ns):
        ax.errorbar(
            n, medians[i],
            yerr=[[yerr_lower[i]], [yerr_upper[i]]],
            fmt="o", color=colors[i],
            capsize=6, capthick=1.8,
            markersize=8, linewidth=1.8,
            label=f"n={n:,}"
        )
    ax.set_title(f"Median + 95% CI at {reference_year}", fontsize=11, fontweight="bold")
    ax.set_xlabel("Sample size (n runs)", fontsize=10)
    ax.set_ylabel("Prevalence", fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    ax.set_xticks(ns)
    ax.set_xticklabels([f"{n:,}" for n in ns], fontsize=9)
    ax.grid(True, alpha=0.3, linewidth=0.5, axis="y")
    ax.spines[["top","right"]].set_visible(False)
    ax.legend(fontsize=9, framealpha=0.7)

    # --- Right: CI width vs sample size ---
    ax2 = axes[1]
    ax2.plot(ns, widths, marker="o", color="#1D3557",
             linewidth=2, markersize=8, zorder=3)
    for i, (n, w) in enumerate(zip(ns, widths)):
        ax2.scatter(n, w, color=colors[i], s=80, zorder=4)
        ax2.annotate(f"{w:,.0f}", (n, w),
                     textcoords="offset points", xytext=(0, 9),
                     ha="center", fontsize=8.5, color=colors[i], fontweight="bold")
    ax2.set_title(f"95% CI Width at {reference_year}", fontsize=11, fontweight="bold")
    ax2.set_xlabel("Sample size (n runs)", fontsize=10)
    ax2.set_ylabel("CI Width (Upper − Lower)", fontsize=10)
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    ax2.set_xticks(ns)
    ax2.set_xticklabels([f"{n:,}" for n in ns], fontsize=9)
    ax2.grid(True, alpha=0.3, linewidth=0.5, axis="y")
    ax2.spines[["top","right"]].set_visible(False)
    ax2.set_ylim(bottom=0)

    plt.tight_layout()
    return fig


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(all_stats: dict, all_focal: dict, scenario_key: str,
                country: str, output_path: Path):

    thin  = Side(style="thin", color="AAAAAA")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="1F4E79")
    sfont = Font(name="Arial", bold=True, size=9)
    cfont = Font(name="Arial", size=9)
    cen   = Alignment(horizontal="center", vertical="center")
    num_fmt = "#,##0.1"

    sample_fills = {n: PatternFill("solid", start_color=c.lstrip("#"))
                    for n, c in {100:"FFF2CC", 250:"FCE4D6",
                                 500:"E2EFDA", 750:"DDEBF7", 1000:"E2E2E2"}.items()}

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Time series ----
    ws1 = wb.create_sheet("Time Series")
    ws1.merge_cells("A1:M1")
    t = ws1["A1"]
    t.value = f"{country} — {outcome_key} | {scenario_key} | Median + 95% CI by sample size"
    t.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    t.alignment = cen

    # sub-headers: Year | [n: median, lower, upper] x5
    ws1.cell(row=2, column=1, value="Year").font = sfont
    ws1.cell(row=2, column=1).alignment = cen
    ws1.cell(row=2, column=1).border = bdr
    ws1.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    col = 2
    col_map = {}
    for n in sample_sizes:
        col_map[n] = col
        ws1.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
        hc = ws1.cell(row=2, column=col, value=f"n = {n:,}")
        hc.font = hfont; hc.fill = hfill; hc.alignment = cen
        for j, sh in enumerate(["Median", "Lower 2.5%", "Upper 97.5%"]):
            c = ws1.cell(row=3, column=col+j, value=sh)
            c.font = sfont; c.fill = sample_fills[n]; c.alignment = cen; c.border = bdr
        col += 4  # 3 data + 1 gap

    # data rows — use years from n=1000
    years = all_stats[1000]["year"].values
    for r, yr in enumerate(years):
        er = r + 4
        yc = ws1.cell(row=er, column=1, value=int(yr) if float(yr).is_integer() else yr)
        yc.font = cfont; yc.alignment = cen; yc.border = bdr
        for n in sample_sizes:
            row_s = all_stats[n][np.isclose(all_stats[n]["year"], yr)]
            if row_s.empty:
                continue
            c = col_map[n]
            for j, field in enumerate(["median","lower","upper"]):
                cell = ws1.cell(row=er, column=c+j, value=round(float(row_s[field].iloc[0]),1))
                cell.font = cfont; cell.alignment = cen; cell.border = bdr
                cell.number_format = num_fmt

    ws1.column_dimensions["A"].width = 8
    for n in sample_sizes:
        c = col_map[n]
        for j in range(3):
            ws1.column_dimensions[get_column_letter(c+j)].width = 13
        ws1.column_dimensions[get_column_letter(c+3)].width = 2

    # ---- Sheet 2: Focal year summary ----
    ws2 = wb.create_sheet(f"{reference_year} Summary")
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = f"{country} — {outcome_key} at {reference_year} | {scenario_key} | Convergence Summary"
    t2.font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    t2.alignment = cen

    sum_headers = ["Sample size (n)", "Median", "Lower 2.5%", "Upper 97.5%",
                   "CI Width", "Converged vs n=1000?"]
    col_ws = [18, 14, 14, 14, 14, 22]
    for ci, (h, w) in enumerate(zip(sum_headers, col_ws), 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = cen; c.border = bdr
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ref_width = all_focal[1000]["ci_width"]
    for ri, n in enumerate(sample_sizes, 3):
        f = all_focal[n]
        if n == 1000:
            conv = "— (reference)"
        else:
            pct_vs_ref = abs(f["ci_width"] - ref_width) / max(ref_width, 1e-9)
            conv = "✓  (<5% vs n=1000)" if pct_vs_ref < 0.05 else f"✗  ({pct_vs_ref*100:.1f}% vs n=1000)"

        row_vals = [n, round(f["median"],1), round(f["lower"],1),
                    round(f["upper"],1), round(f["ci_width"],1), conv]
        for ci, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = cfont; cell.alignment = cen; cell.border = bdr
            cell.fill = sample_fills.get(n, PatternFill())
            if ci in (2,3,4,5):
                cell.number_format = num_fmt
            if ci == 6 and val:
                if "reference" in str(val):
                    cell.font = Font(name="Arial", size=9, color="666666", italic=True)
                else:
                    cell.font = Font(name="Arial", size=9, bold=True,
                                     color="375623" if "✓" in str(val) else "C00000")

    wb.save(output_path)
    print(f"  Excel saved: {output_path}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def convergence_check(pkl_path: str | Path,
                      output_dir: str | Path = None,
                      scenarios: list = None) -> None:
    """
    Parameters
    ----------
    pkl_path   : path to *_econ_eval.pkl
    output_dir : folder for outputs (default: same folder as pkl)
    scenarios  : list of scenario keys to analyse (default: all)
    """
    pkl_path = Path(pkl_path)
    country  = pkl_path.stem.replace("_econ_eval", "")
    out_dir  = Path(output_dir) if output_dir else pkl_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading {pkl_path.name} ...")
    country_data  = sc.load(pkl_path)
    agg           = country_data["agg_data"]
    scenario_keys = scenarios or list(agg.keys())
    print(f"  Scenarios: {scenario_keys}")

    for scen_key in scenario_keys:
        print(f"\n  Processing scenario: {scen_key}")
        if outcome_key not in agg[scen_key]:
            print(f"    '{outcome_key}' not found — skipping")
            continue

        df_outcome = agg[scen_key][outcome_key]

        # Compute stats for each sample size
        all_stats = {n: sample_stats(df_outcome, n) for n in sample_sizes}
        all_focal = {n: focal_year_stats(all_stats[n]) for n in sample_sizes}

        # Figure 1: time series grid
        fig1 = plot_timeseries_grid(all_stats, scen_key, country)
        p1 = out_dir / f"{country}_{scen_key}_convergence_timeseries.png"
        fig1.savefig(p1, dpi=150, bbox_inches="tight")
        plt.close(fig1)
        print(f"  Figure 1 saved: {p1.name}")

        # Figure 2: focal year CI
        fig2 = plot_focal_year(all_focal, scen_key, country)
        p2 = out_dir / f"{country}_{scen_key}_convergence_{reference_year}.png"
        fig2.savefig(p2, dpi=150, bbox_inches="tight")
        plt.close(fig2)
        print(f"  Figure 2 saved: {p2.name}")

        # Excel
        px = out_dir / f"{country}_{scen_key}_convergence.xlsx"
        write_excel(all_stats, all_focal, scen_key, country, px)


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python convergence_check.py <pkl_path> [output_dir] [scen1,scen2,...]")
        sys.exit(1)
    pkl     = Path(sys.argv[1])
    out     = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    scens   = sys.argv[3].split(",") if len(sys.argv) > 3 else None
    convergence_check(pkl, out, scens)